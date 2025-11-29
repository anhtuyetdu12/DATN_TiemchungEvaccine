import re
from datetime import datetime, date as date_cls
from rest_framework import viewsets, permissions, status, generics
from django.utils.html import strip_tags, escape
import openpyxl
from django.core.mail import EmailMultiAlternatives
from email.mime.image import MIMEImage
from pathlib import Path
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from vaccines.models import Vaccine
from django.db import transaction
from django.utils import timezone
from django.db.models import Q, Sum
from datetime import timedelta
from django.utils.dateparse import parse_date
from inventory.models import VaccineStockLot, BookingAllocation
from users.models import CustomUser 
from .models import FamilyMember, VaccinationRecord, Booking,  CustomerNotification
from .serializers import ( 
    FamilyMemberSerializer, VaccinationRecordSerializer,                      
    CustomerListSerializer, CustomerMemberSlimSerializer,
    AppointmentStatusPatchSerializer, CustomerNotificationSerializer,
    HistoryCreateInSerializer, StaffBookingCreateInSerializer, BookingSerializer
)
from django.core.mail import send_mail
from django.conf import settings

class FamilyMemberViewSet(viewsets.ModelViewSet):
    serializer_class = FamilyMemberSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return FamilyMember.objects.filter(user=self.request.user)
    
    def perform_create(self, serializer):
        # GÁN user cho bản ghi mới => tránh NULL user_id
        serializer.save(user=self.request.user)

    def perform_update(self, serializer):
        # (tuỳ chọn) đảm bảo không đổi chủ sở hữu
        serializer.save(user=self.request.user)

    def list(self, request, *args, **kwargs):
        user = request.user
        queryset = self.get_queryset()
        # Kiểm tra đã có "Bản thân" chưa
        self_member = queryset.filter(relation="Bản thân").first()
        if not self_member:
            # Tạo default member cho user chính
            self_member = FamilyMember.objects.create(
                user=user,
                full_name=getattr(user, "full_name", "") or user.email,
                nickname=getattr(user, "full_name", "") or user.email,
                relation="Bản thân",
                gender="other",
                date_of_birth=date_cls(2000, 1, 1),
                phone=getattr(user, "phone", ""),
                is_self=True  # nếu model có cột này
            )
            # Đặt queryset gồm member mới + các thành viên khác
            other_members = queryset.exclude(id=self_member.id)
            queryset = [self_member] + list(other_members)
        else:
            # Đảm bảo “Bản thân” luôn lên đầu danh sách
            other_members = queryset.exclude(id=self_member.id)
            queryset = [self_member] + list(other_members)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    

# --- Quản lý sổ tiêm chủng ---
class VaccinationRecordViewSet(viewsets.ModelViewSet):
    queryset = VaccinationRecord.objects.all()  
    serializer_class = VaccinationRecordSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        member_id = self.request.query_params.get("member_id")
        queryset = VaccinationRecord.objects.filter(
            family_member__user=self.request.user
        ).filter(
            Q(vaccination_date__isnull=False) | Q(next_dose_date__isnull=False)
        )
        if member_id:
            queryset = queryset.filter(family_member_id=member_id)
        return queryset.order_by("-next_dose_date", "-vaccination_date")
    
    def perform_create(self, serializer):
        serializer.save()

# ---------- đặt hẹn mũi tiêm  -----------
class RemainingDosesView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        member_id = request.query_params.get("member_id")
        vaccine_id = request.query_params.get("vaccine_id")
        if not member_id or not vaccine_id:
            return Response({"error": "Thiếu member_id/vaccine_id"}, status=400)
        role = getattr(request.user, "role", "")
        if role in ("staff", "admin", "superadmin"):
            member = FamilyMember.objects.filter(id=member_id).first()
        else:
            # customer bình thường -> chỉ được xem member của chính mình
            member = FamilyMember.objects.filter(
                id=member_id,
                user=request.user
            ).first()
        vaccine = Vaccine.objects.filter(id=vaccine_id).first()
        if not member or not vaccine:
            return Response({"error": "Không hợp lệ"}, status=400)
        # Tổng mũi theo phác đồ (doses_required)
        total = vaccine.doses_required or 1
        # Đã TIÊM bao nhiêu mũi (có vaccination_date)
        used = VaccinationRecord.objects.filter(
            family_member=member,
            vaccine=vaccine,
            vaccination_date__isnull=False,
        ).count()
        remaining = max(total - used, 0)
        # Tìm record "mũi kế tiếp" (đã được sinh trong Booking.complete / StaffUpdateAppointmentStatusAPIView)
        next_rec = (
            VaccinationRecord.objects
            .filter(
                family_member=member,
                vaccine=vaccine,
                vaccination_date__isnull=True,
                next_dose_date__isnull=False,
            )
            .order_by("next_dose_date")
            .first()
        )
        next_date = next_rec.next_dose_date if next_rec else None
        # Nếu đã có record dự kiến thì dùng luôn dose_number từ đó
        if next_rec and next_rec.dose_number:
            next_dose_number = next_rec.dose_number
        else:
            next_dose_number = used + 1 if remaining > 0 else None
        # Nếu CHƯA có record dự kiến mà vẫn còn mũi
        if not next_date and remaining > 0:
            interval = getattr(vaccine, "interval_days", None)
            today = timezone.now().date()
            last_rec = (
                VaccinationRecord.objects
                .filter(
                    family_member=member,
                    vaccine=vaccine,
                    vaccination_date__isnull=False,
                )
                .order_by("-vaccination_date")
                .first()
            )
            if last_rec and last_rec.vaccination_date and interval:
                # đã tiêm ≥1 mũi 
                next_date = last_rec.vaccination_date + timedelta(days=interval)
            elif used == 0:
                # chưa tiêm mũi nào -> gợi ý có thể bắt đầu từ hôm nay
                next_date = today
        # Trạng thái phác đồ
        if used >= total:
            status_code = "completed"
            status_label = f"Đã tiêm đủ {used}/{total} mũi theo phác đồ."
        elif used == 0:
            status_code = "not_started"
            status_label = f"Chưa tiêm mũi nào. Phác đồ gồm {total} mũi."
        else:
            status_code = "in_progress"
            status_label = f"Đã tiêm {used}/{total} mũi. Còn {total - used} mũi."

        if next_date and status_code != "completed":
            status_label += (
                f" Mũi {next_dose_number} dự kiến từ ngày "
                f"{next_date.strftime('%d/%m/%Y')}."
            )
        return Response({
            "total": total,
            "used": used,
            "remaining": remaining,
            "status_code": status_code,          # not_started | in_progress | completed
            "status_label": status_label,        # Chuỗi hiển thị cho khách
            "next_dose_date": next_date,         # ISO string khi serialize
            "next_dose_number": next_dose_number,
        })
    
    
# ---------- booking -----------
class BookingViewSet(viewsets.ModelViewSet):
    queryset = Booking.objects.all().select_related("user", "member").prefetch_related("items__vaccine")
    serializer_class = BookingSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        role = getattr(user, "role", "")
        
        if not user.is_superuser and role not in ("staff", "admin", "superadmin"):
            qs = qs.filter(user=user)
        status_param = self.request.query_params.get("status")
        search = self.request.query_params.get("search")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        if status_param == "overdue":
            today = timezone.now().date()
            qs = qs.filter(
                appointment_date__lt=today
            ).exclude(status__in=["completed", "cancelled"])
        elif status_param == "pending":
            today = timezone.now().date()
            qs = qs.filter(
                status="pending",
                appointment_date__gte=today,
            )
        elif status_param in ("confirmed", "completed", "cancelled"):
            qs = qs.filter(status=status_param)

        if search:
            qs = qs.filter(
                Q(user__email__icontains=search) |
                Q(member__full_name__icontains=search) |
                Q(location__icontains=search)
            )
        if date_from:
            d1 = parse_date(date_from)
            if d1:
                qs = qs.filter(appointment_date__gte=d1)
        if date_to:
            d2 = parse_date(date_to)
            if d2:
                qs = qs.filter(appointment_date__lte=d2)
        return qs.order_by("-appointment_date", "-created_at")

    @action(detail=True, methods=["POST"], url_path="confirm")
    def confirm(self, request, pk=None):
        booking = self.get_object()
        if booking.status in ("cancelled", "completed"):
            return Response({"detail": "Không thể xác nhận lịch đã hủy/đã hoàn thành."}, status=400)
        if booking.status == "confirmed":
            return Response({"status": "confirmed"}, status=200)
        try:
            with transaction.atomic():
                for item in booking.items.select_related("vaccine").all():
                    if not item.vaccine:
                        continue
                    need = int(item.quantity or 0)
                    if need <= 0:
                        continue
                    today = timezone.now().date()
                    appt_date = booking.appointment_date or today
                    # TẤT CẢ CÁC LÔ ĐANG ACTIVE (kể cả hết hạn) để biết còn hàng hay không
                    all_qs = VaccineStockLot.objects.filter(
                        vaccine=item.vaccine,
                        is_active=True,
                        quantity_available__gt=0,
                    )
                    total_all = all_qs.aggregate(total=Sum("quantity_available"))["total"] or 0
                    # CÁC LÔ CÒN HẠN tính theo NGÀY HẸN TIÊM
                    usable_qs = all_qs.filter(expiry_date__gte=appt_date)
                    usable_total = usable_qs.aggregate(total=Sum("quantity_available"))["total"] or 0
                    # 1) Không còn hàng (dù hạn hay hết hạn)
                    if total_all == 0:
                        raise ValueError( f"Vắc xin {item.vaccine.name} hiện đã hết số lượng trong kho." )
                    # 2) Còn hàng nhưng tất cả đều hết hạn
                    if usable_total == 0:
                        raise ValueError(
                            f"Tất cả các lô của vắc xin {item.vaccine.name} "
                            f"đã hết hạn sử dụng trước ngày hẹn tiêm. "
                            f"Vui lòng nhập lô mới hoặc đổi vắc xin khác."
                        )
                    # 3) Còn vắc xin còn hạn nhưng KHÔNG ĐỦ cho số lượng đặt
                    if usable_total < need:
                        raise ValueError(
                            f"Tồn kho không đủ cho {item.vaccine.name}. "
                            f"Hiện còn {usable_total} liều (còn hạn), cần {need} liều."
                        )
                    # 4) ĐỦ HÀNG → trừ tồn kho theo FIFO CHỈ TRÊN CÁC LÔ CÒN HẠN
                    lots = (
                        usable_qs.select_for_update()
                        .order_by("expiry_date")
                    )
                    for lot in lots:
                        if need <= 0:
                            break
                        take = min(lot.quantity_available, need)
                        if take > 0:
                            lot.quantity_available -= take
                            lot.save(update_fields=["quantity_available"])
                            BookingAllocation.objects.create(
                                booking_item=item,
                                lot=lot,
                                quantity=take,
                                status="reserved",
                            )
                            need -= take
                    if need > 0:
                        # Về lý thuyết không vào đây, nhưng để an toàn
                        raise ValueError( f"Tồn kho không đủ cho {item.vaccine.name} sau khi trừ lô. Thiếu {need} liều.")
                booking.status = "confirmed"
                booking.save(update_fields=["status"])
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)
        return Response({"status": "confirmed"})
    
    @action(detail=True, methods=["POST"], url_path="cancel")
    def cancel(self, request, pk=None):
        booking = self.get_object()
        if booking.status == "completed":
            return Response({"detail":"Không thể hủy lịch đã tiêm."}, status=400)

        with transaction.atomic():
            # 1) Trả hàng từ các allocation đang giữ
            for item in booking.items.all():
                for alloc in item.allocations.select_for_update().filter(status="reserved"):
                    lot = alloc.lot
                    lot.quantity_available += alloc.quantity
                    lot.save(update_fields=["quantity_available"])
                    alloc.status = "released"
                    alloc.save(update_fields=["status"])
            # 2) XÓA next_dose_date của các dòng sổ tiêm “dự kiến” thuộc lịch này
            from records.models import VaccinationRecord
            VaccinationRecord.objects.filter(
                family_member=booking.member,
                vaccination_date__isnull=True,
            ).filter(
                Q(source_booking=booking) | Q(next_dose_date=booking.appointment_date)
            ).update(next_dose_date=None)
        booking.status = "cancelled"
        booking.save(update_fields=["status"])
        return Response({"status":"cancelled"})

    @action(detail=True, methods=["POST"], url_path="complete")
    def complete(self, request, pk=None):
        booking = self.get_object()
        if booking.status == "cancelled":
            return Response({"detail":"Lịch đã hủy không thể hoàn thành."}, status=400)
        if booking.status == "completed":
            return Response({"detail":"Lịch đã hoàn thành."}, status=200)
        reaction_note = (request.data.get("reaction_note") or "").strip()

        with transaction.atomic():
            # (1) Đổi allocation -> consumed
            for item in booking.items.all():
                item.allocations.filter(status="reserved").update(status="consumed")
            today = timezone.now().date()
            # (2) Đánh dấu các record dự kiến tương ứng là đã tiêm hôm nay
            rec_qs = VaccinationRecord.objects.select_for_update().filter(
                family_member=booking.member,
                vaccine__in=booking.items.values("vaccine_id"),
                next_dose_date=booking.appointment_date,
                vaccination_date__isnull=True,
            )
            updated = 0
            for rec in rec_qs:
                rec.vaccination_date = today
                rec.next_dose_date = None
                if reaction_note:
                    rec.note = (rec.note + "\n" if rec.note else "") + reaction_note
                rec.save(update_fields=["vaccination_date", "next_dose_date", "note"])
                updated += 1
            # (3) Sinh mũi kế tiếp theo phác đồ (nếu còn)
            for item in booking.items.all():
                v = item.vaccine
                if not v:
                    continue
                total = getattr(v, "doses_required", 1) or 1
                interval = getattr(v, "interval_days", None)
                taken = VaccinationRecord.objects.filter(
                    family_member=booking.member,
                    vaccine=v,
                    vaccination_date__isnull=False,
                ).count()
                if taken >= total:
                    continue
                if not interval:
                    continue
                next_date = today + timedelta(days=interval)
                VaccinationRecord.objects.create(
                    family_member=booking.member,
                    disease=v.disease,
                    vaccine=v,
                    dose_number=taken + 1,
                    vaccination_date=None,
                    next_dose_date=next_date,
                    note=f"Tự sinh mũi kế tiếp từ booking #{booking.id}",
                )
            # (4) GÁN SỐ LÔ VÀO VaccinationRecord  # NEW
            for item in booking.items.select_related("vaccine").all():
                v = item.vaccine
                if not v:
                    continue
                # Các allocation đã dùng cho item này
                allocs = (
                    item.allocations
                    .filter(status="consumed")
                    .select_related("lot")
                    .order_by("id")
                )
                if not allocs.exists():
                    continue
                # Các record tương ứng mũi đã tiêm hôm nay, chưa có số lô
                recs = list(
                    VaccinationRecord.objects
                    .filter(
                        family_member=booking.member,
                        vaccine=v,
                        vaccination_date=today,
                    )
                    .filter(Q(vaccine_lot__isnull=True) | Q(vaccine_lot__exact=""))
                    .order_by("id")
                )
                if not recs:
                    continue
                # quantity hiện tại của bạn đang giới hạn =1/mũi, nhưng làm general: map từng alloc -> từng rec
                rec_iter = iter(recs)
                for alloc in allocs:
                    lot = alloc.lot
                    if not lot or not lot.lot_number:
                        continue
                    try:
                        rec = next(rec_iter)
                    except StopIteration:
                        break
                    rec.vaccine_lot = lot.lot_number
                    rec.save(update_fields=["vaccine_lot"])
            # (5) Hoàn tất booking
            booking.status = "completed"
            booking.save(update_fields=["status"])
        return Response({"status": "completed", "updated_records": updated}, status=200)
    
    @action(detail=False, methods=["GET"], url_path="export/excel")
    def export_excel(self, request):
        # Lấy cùng queryset với trang quản lý (đã áp dụng status, search, date_from, date_to, role...)
        qs = self.get_queryset()
        # Không phân trang ở đây: xuất full kết quả theo filter
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lịch hẹn"

        # Header
        headers = [
            "ID",
            "Email khách",
            "Người tiêm",
            "Ngày hẹn",
            "Vắc xin / Gói",
            "Trạng thái",
            "Tổng tiền (VNĐ)",
            "Cơ sở",
        ]
        ws.append(headers)
        for b in qs:
            # Tính tổng tiền + tên vắc xin giống StaffCustomerListAPIView
            item_names = []
            total_price = 0
            for it in b.items.all():
                if it.vaccine:
                    q = int(it.quantity or 0)
                    item_names.append(f"{it.vaccine.name} x{q}")
                    try:
                        total_price += int(float(it.unit_price or 0)) * q
                    except Exception:
                        pass
            vaccine_label = ", ".join(item_names) or ( b.vaccine.name if b.vaccine else (f"Gói: {b.package.name}" if b.package else ""))
            # status_label dùng luôn helper trong BookingSerializer
            ser = BookingSerializer(b, context={"request": request})
            status_label = ser.data.get("status_label", b.status)
            ws.append([
                b.id,
                b.user.email if b.user else "",
                b.member.full_name if b.member else "",
                b.appointment_date.strftime("%d/%m/%Y") if b.appointment_date else "",
                vaccine_label,
                status_label,
                total_price,
                b.location or "Trung tâm tiêm chủng E-Vaccine",
            ])
        # Auto-width đơn giản
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2
        # Trả file về client
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = "lich-hen.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


# ---------- Trả danh sách “khách hàng”  -----------
class StaffCustomerListAPIView(APIView):
    """
    GET /api/records/staff/customers/?include=members
    Trả đúng shape mà FE đang dùng (CustomerListSerializer)
    """
    permission_classes = [IsAuthenticated]
    def get(self, request):
        include = request.query_params.get("include", "")
        include_members = "members" in [s.strip() for s in include.split(",") if s]

        qs = CustomUser.objects.filter(role="customer")
        role = getattr(request.user, "role", "")
        if role not in ("staff", "admin", "superadmin"):
            qs = qs.filter(id=request.user.id)

        data = []
        for u in qs:
            code = f"KH-{u.id:04d}"
            # LẤY MEMBER “BẢN THÂN” CHO USER HIỆN TẠI
            m_self = (
                FamilyMember.objects
                .filter(user=u, relation="Bản thân")
                .order_by("-is_self", "-created_at")
                .first()
            )
            # 5 lịch hẹn gần nhất
            bookings = (
                Booking.objects.filter(user=u)
                .select_related("vaccine", "package")
                .prefetch_related("items__vaccine")
                .order_by("-appointment_date")[:5]
            )
            appts = []
            for b in bookings:
                item_names = []
                total_price = 0
                for it in b.items.all():
                    if it.vaccine:
                        q = int(it.quantity or 0)
                        item_names.append(f"{it.vaccine.name} x{q}")
                        try:
                            total_price += int(float(it.unit_price or 0)) * q
                        except:
                            pass
                vaccine_label = ", ".join(item_names) or (
                    b.vaccine.name if b.vaccine else (f"Gói: {b.package.name}" if b.package else "")
                )
                appts.append({
                    "id": str(b.id),
                    "date": b.appointment_date,
                    "vaccine": vaccine_label,
                    "center": b.location or "",
                    "status": b.status,
                    "price": int(total_price),
                })
            # lịch sử tiêm (tối đa 50)
            recs = (
                VaccinationRecord.objects
                .filter(family_member__user=u)
                .select_related("family_member", "vaccine", "disease")
                .exclude(vaccination_date__isnull=True)
                .order_by("-vaccination_date")[:50]
            )
            history = []
            for r in recs:
                history.append({
                    "id": str(r.id),
                    "date": r.vaccination_date,
                    "member_id": r.family_member_id,
                    "member_name": r.family_member.full_name if r.family_member else "",
                    "relation": r.family_member.relation if r.family_member else "",
                    "disease": r.disease.name if r.disease else "",
                    "vaccine": r.vaccine.name if r.vaccine else (r.vaccine_name or ""),
                    "dose": int(r.dose_number or 1),
                    "price": int(getattr(r.vaccine, "price", 0) or 0),
                    "note": r.note or "",
                    "batch": r.vaccine_lot or "",   
                    "status_label": "Đã tiêm",
                })
            doses_count = (
                VaccinationRecord.objects
                .filter(family_member__user=u)
                .exclude(vaccination_date__isnull=True)
                .count()
            )
            row = {
                "id": int(u.id),
                "code": code,
                "name": u.full_name or u.email,
                "phone": u.phone or "",
                "email": u.email,
                "dob": m_self.date_of_birth if m_self else None, 
                "gender": (m_self.gender or "") if m_self else "",
                "address": "",
                "country": "Việt Nam",
                "category": "",
                "doses": int(doses_count),
                "appointments": appts,
                "history": history,
            }
            if include_members:
                members = FamilyMember.objects.filter(user=u).order_by("-is_self", "-created_at")
                row["members"] = [
                    {
                        "id": m.id,
                        "full_name": m.full_name,
                        "nickname": m.nickname or "",
                        "relation": m.relation or "",
                        "gender": m.gender or "",
                        "date_of_birth": m.date_of_birth,
                        "phone": getattr(m, "phone", "") or "",
                    } for m in members
                ]
            data.append(row)
        ser = CustomerListSerializer(data, many=True)
        return Response(ser.data, status=200)

# ---------- lấy members của 1 user cụ thể.  -----------
class StaffCustomerMembersAPIView(APIView):
    """
    GET /api/records/staff/customers/<user_id>/members/
    """
    permission_classes = [IsAuthenticated]
    def get(self, request, user_id):
        role = getattr(request.user, "role", "")
        if role not in ("staff", "admin", "superadmin") and request.user.id != user_id:
            return Response({"detail": "Forbidden"}, status=403)
        members = FamilyMember.objects.filter(user_id=user_id).order_by("-created_at")
        ser = CustomerMemberSlimSerializer(members, many=True)
        return Response(ser.data, status=200)

# ---------- Lịch hẹn tiêm chủng  -----------
class StaffListAppointmentsAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, user_id: int):
        role = getattr(request.user, "role", "")
        if role not in ("staff", "admin", "superadmin") and request.user.id != user_id:
            return Response({"detail": "Forbidden"}, status=403)
        bookings = (
            Booking.objects.filter(user_id=user_id)
            .select_related("vaccine", "package", "member")
            .prefetch_related("items__vaccine__disease")
            .order_by("-appointment_date")
        )

        # (tuỳ chọn) giữ nguyên lọc theo ?status
        status_q = request.query_params.get("status", "")
        if status_q:
            allowed = {"pending", "confirmed", "completed", "cancelled"}
            want = [s.strip() for s in status_q.split(",") if s.strip() in allowed]
            if want:
                bookings = bookings.filter(status__in=want)
        today = timezone.now().date()
        data = []
        for b in bookings:
            items_summary, disease_names, total_price = [], [], 0
            for it in b.items.all():
                if it.vaccine:
                    q = int(it.quantity or 0)
                    items_summary.append({"name": it.vaccine.name, "qty": q})
                    if it.vaccine.disease:
                        disease_names.append(it.vaccine.disease.name)
                    try:
                        total_price += int(float(it.unit_price or 0)) * q
                    except:
                        pass
            vaccine_label = ", ".join([f"{s['name']} x{s['qty']}" for s in items_summary]) or (
                b.vaccine.name if b.vaccine else (f"Gói: {b.package.name}" if b.package else "")
            )
            # ----- NEW: tính trễ hẹn -----
            is_overdue = (
                b.status not in ("completed", "cancelled")
                and b.appointment_date is not None
                and b.appointment_date < today
            )
            effective_status = "overdue" if is_overdue else b.status

            planned = VaccinationRecord.objects.filter(
                family_member=b.member,
                next_dose_date=b.appointment_date,
            ).order_by("dose_number")
            doses = [p.dose_number for p in planned if p.dose_number is not None]
            dose = doses[0] if len(doses) == 1 else None

            status_label = (
                "Trễ hẹn" if is_overdue else {
                    "pending": "Chờ xác nhận",
                    "confirmed": "Đã xác nhận",
                    "completed": "Đã tiêm xong",
                    "cancelled": "Đã hủy",
                }.get(b.status, b.status)
            )

            data.append({
                "id": str(b.id),
                "date": b.appointment_date,
                "vaccine": vaccine_label,
                "items_summary": items_summary,
                "center": b.location or "",
                "status": b.status, 
                "status_label": status_label,
                "effective_status": effective_status,
                "is_overdue": is_overdue,      
                "price": int(total_price),
                "member_id": b.member_id,
                "member_name": b.member.full_name if b.member else "",
                "relation": b.member.relation if b.member else "",
                "disease": ", ".join(dict.fromkeys(disease_names)) if disease_names else "",
                "dose": dose,
            })
        return Response(data, status=200)


# ---------- cập nhật trạng thái - patch -----------
class StaffUpdateAppointmentStatusAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def patch(self, request, user_id: int, appt_id: str):
        role = getattr(request.user, "role", "")
        if role not in ("staff", "admin", "superadmin") and request.user.id != user_id:
            return Response({"detail": "Forbidden"}, status=403)
        ser = AppointmentStatusPatchSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        new_status = ser.validated_data["status"]
        booking = Booking.objects.filter(id=appt_id, user_id=user_id).first()
        if not booking:
            return Response({"detail": "Lịch hẹn không tồn tại"}, status=404)
        if booking.status == "completed" and new_status != "completed":
            return Response({"detail": "Không thể đổi trạng thái lịch đã hoàn thành"}, status=400)

        with transaction.atomic():
            if new_status == "cancelled":
                # trả hàng & gỡ lịch dự kiến, giống BookingViewSet.cancel
                for item in booking.items.all():
                    for alloc in item.allocations.select_for_update().filter(status="reserved"):
                        lot = alloc.lot
                        lot.quantity_available += alloc.quantity
                        lot.save(update_fields=["quantity_available"])
                        alloc.status = "released"
                        alloc.save(update_fields=["status"])
                    VaccinationRecord.objects.filter(
                        family_member=booking.member,
                        vaccination_date__isnull=True,
                    ).filter(
                        Q(source_booking=booking) | Q(next_dose_date=booking.appointment_date)
                    ).update(next_dose_date=None)

            elif new_status == "completed":
                from records.models import VaccinationRecord
                today = timezone.now().date()
                # 1) Đổi allocation reserved -> consumed
                for item in booking.items.all():
                    item.allocations.filter(status="reserved").update(status="consumed")
                # 2) Cập nhật các record dự kiến (nếu có)
                planned_qs = VaccinationRecord.objects.select_for_update().filter(
                    family_member=booking.member,
                    vaccine__in=booking.items.values("vaccine_id"),
                    next_dose_date=booking.appointment_date,
                    vaccination_date__isnull=True,
                )
                planned_qs.update(vaccination_date=today, next_dose_date=None)
                # 3) Với từng vaccine trong booking: đảm bảo có VaccinationRecord cho hôm nay
                for item in booking.items.select_related("vaccine").all():
                    v = item.vaccine
                    if not v:
                        continue
                    # Các allocations đã dùng (đã consumed)
                    allocs = list(
                        item.allocations
                        .filter(status="consumed")
                        .select_related("lot")
                        .order_by("id")
                    )
                    if not allocs:
                        continue
                    # Tìm các record đã có cho hôm nay, chưa có số lô
                    recs = list(
                        VaccinationRecord.objects
                        .filter(
                            family_member=booking.member,
                            vaccine=v,
                            vaccination_date=today,
                        )
                        .order_by("id")
                    )
                    # Nếu chưa có record nào cho hôm nay -> tạo mới tương ứng số liều đã dùng
                    if not recs:
                        # Đếm số mũi đã tiêm trước đó cho vaccine này
                        taken_before = VaccinationRecord.objects.filter(
                            family_member=booking.member,
                            vaccine=v,
                            vaccination_date__lt=today,
                        ).count()
                        # Mỗi allocation.quantity ~ số liều; ở đây phần lớn là 1
                        dose_number = taken_before + 1
                        for alloc in allocs:
                            lot = alloc.lot
                            VaccinationRecord.objects.create(
                                family_member=booking.member,
                                disease=v.disease,
                                vaccine=v,
                                dose_number=dose_number,
                                vaccination_date=today,
                                next_dose_date=None,
                                vaccine_lot=(lot.lot_number if lot else None),
                                note=f"Tự tạo từ lịch #{booking.id}",
                            )
                            dose_number += int(alloc.quantity or 1)
                        # Sau khi create xong thì sang item tiếp theo
                        continue
                    # Nếu đã có recs (từ planned_qs), thì map số lô vào các rec chưa có vaccine_lot
                    empty_lot_recs = [r for r in recs if not (r.vaccine_lot or "").strip()]
                    if not empty_lot_recs:
                        continue
                    rec_iter = iter(empty_lot_recs)
                    for alloc in allocs:
                        lot = alloc.lot
                        if not lot or not lot.lot_number:
                            continue
                        try:
                            rec = next(rec_iter)
                        except StopIteration:
                            break
                        rec.vaccine_lot = lot.lot_number
                        rec.save(update_fields=["vaccine_lot"])
                # 4) Sinh mũi kế tiếp theo phác đồ (nếu còn)
                for item in booking.items.all():
                    v = item.vaccine
                    if not v:
                        continue
                    total = getattr(v, "doses_required", 1) or 1
                    interval = getattr(v, "interval_days", None)
                    taken = VaccinationRecord.objects.filter(
                        family_member=booking.member,
                        vaccine=v,
                        vaccination_date__isnull=False,
                    ).count()
                    if taken >= total or not interval:
                        continue
                    next_date = today + timedelta(days=interval)
                    VaccinationRecord.objects.create(
                        family_member=booking.member,
                        disease=v.disease,
                        vaccine=v,
                        dose_number=taken + 1,
                        vaccination_date=None,
                        next_dose_date=next_date,
                        note=f"Tự sinh mũi kế tiếp từ booking #{booking.id}",
                    )
            booking.status = new_status
            booking.save(update_fields=["status"])
            booking = ( Booking.objects.select_related("vaccine", "package")
                .prefetch_related("items__vaccine")
                .get(id=booking.id) )
            total_price = sum([(int(float(it.unit_price or 0))) * int(it.quantity or 0) for it in booking.items.all()])
            names = [f"{it.vaccine.name} x{int(it.quantity or 0)}" for it in booking.items.all() if it.vaccine]
            vaccine_label = ", ".join(names) or (booking.vaccine.name if booking.vaccine else (f"Gói: {booking.package.name}" if booking.package else ""))
            return Response({
                "id": str(booking.id),
                "date": booking.appointment_date,
                "vaccine": vaccine_label,
                "center": booking.location or "",
                "status": booking.status,
                "price": int(total_price),
            }, status=200)

# ---------- Lịch sử tiêm chủng -----------
class StaffAddHistoryAPIView(APIView):
    """ POST /api/records/staff/customers/<user_id>/history """
    permission_classes = [IsAuthenticated]
    
    def post(self, request, user_id: int):
        role = getattr(request.user, "role", "")
        if role not in ("staff", "admin", "superadmin") and request.user.id != user_id:
            return Response({"detail": "Forbidden"}, status=403)
        ser = HistoryCreateInSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        data = ser.validated_data
        user = CustomUser.objects.filter(id=user_id, role="customer").first()
        if not user:
            return Response({"detail": "Customer không tồn tại"}, status=404)
        member = None
        mid = data.get("member_id")

        if mid:
            # chỉ chấp nhận member thuộc user này
            member = FamilyMember.objects.filter(id=mid, user=user).first()
            if not member:
                return Response({"detail": "Thành viên không thuộc tài khoản này"}, status=400)

        # nếu không gửi member_id → fallback về “Bản thân” như cũ
        if not member:
            member = FamilyMember.objects.filter(user=user).order_by("-is_self", "-created_at").first()
        if not member:
            return Response({"detail": "Chưa có hồ sơ thành viên"}, status=400)
        rec = VaccinationRecord.objects.create(
            family_member=member,
            vaccine=None,
            vaccine_name=data["vaccine"],
            vaccine_lot=data.get("batch") or "",
            vaccination_date=data["date"],
            next_dose_date=None,
            note=data.get("note") or "",
        )
        return Response({
            "id": str(rec.id),
            "date": rec.vaccination_date,
            "member_id": rec.family_member_id,
            "member_name": rec.family_member.full_name if rec.family_member else "",
            "vaccine": rec.vaccine.name if rec.vaccine else (rec.vaccine_name or ""),
            "batch": rec.vaccine_lot or "",
            "note": rec.note or "",
        }, status=201)
        
# -----------staff cập nhật tt ng dùng------------------
class StaffUpdateCustomerProfileAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, user_id: int):
        role = getattr(request.user, "role", "")
        if role not in ("staff", "admin", "superadmin"):
            return Response({"detail": "Forbidden"}, status=403)
        user = CustomUser.objects.filter(id=user_id, role="customer").first()
        if not user:
            return Response({"detail": "Customer không tồn tại"}, status=404)

        full_name = (request.data.get("full_name") or "").strip()
        phone     = (request.data.get("phone") or "").strip()
        dob       = request.data.get("date_of_birth")  # "YYYY-MM-DD"
        gender    = (request.data.get("gender") or "").strip()  # male|female|other

        if full_name:
            user.full_name = full_name
        if phone:
            user.phone = phone
        user.save(update_fields=["full_name", "phone"])

        m_self = FamilyMember.objects.filter(user=user, relation="Bản thân").first()
        if not m_self:
            m_self = FamilyMember.objects.create(
                user=user,
                full_name=user.full_name or user.email,
                nickname=user.full_name or user.email,
                relation="Bản thân",
                gender="other",
                date_of_birth=date_cls(2000,1,1),
                is_self=True
            )
        if dob:
            m_self.date_of_birth = dob
        if gender in ("male","female","other"):
            m_self.gender = gender
        m_self.full_name = user.full_name or m_self.full_name
        m_self.phone = user.phone or m_self.phone
        m_self.save()
        return Response({
            "user": {
                "id": user.id,
                "full_name": user.full_name,
                "phone": user.phone,
            },
            "self_member": {
                "id": m_self.id,
                "date_of_birth": m_self.date_of_birth,
                "gender": m_self.gender,
            }
        }, status=200)
        
#-----------  staff qly thành viên -----------------
class StaffManageMemberAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, user_id: int):
        role = getattr(request.user, "role", "")
        if role not in ("staff", "admin", "superadmin"):
            return Response({"detail": "Forbidden"}, status=403)
        user = CustomUser.objects.filter(id=user_id, role="customer").first()
        if not user:
            return Response({"detail": "Customer không tồn tại"}, status=404)

        payload = request.data or {}
        fm = FamilyMember.objects.create(
            user=user,
            full_name=payload.get("full_name") or payload.get("name") or "",
            nickname=payload.get("nickname") or payload.get("full_name") or "",
            relation=payload.get("relation") or "Khác",
            gender=payload.get("gender") or "other",
            date_of_birth=payload.get("date_of_birth"),
            phone=payload.get("phone") or "",
            is_self=False,
        )
        return Response({
            "id": fm.id,
            "full_name": fm.full_name,
            "relation": fm.relation,
            "gender": fm.gender,
            "date_of_birth": fm.date_of_birth,
            "phone": fm.phone,
        }, status=201)

    def patch(self, request, user_id: int, member_id: int):
        role = getattr(request.user, "role", "")
        if role not in ("staff", "admin", "superadmin"):
            return Response({"detail": "Forbidden"}, status=403)
        fm = FamilyMember.objects.filter(id=member_id, user_id=user_id).first()
        if not fm:
            return Response({"detail": "Thành viên không tồn tại"}, status=404)
        data = request.data or {}
        for f in ["full_name", "nickname", "relation", "gender", "date_of_birth", "phone"]:
            if f in data and data[f] is not None:
                setattr(fm, f, data[f])
        fm.save()
        return Response({"detail": "Updated"}, status=200)

    def delete(self, request, user_id: int, member_id: int):
        role = getattr(request.user, "role", "")
        if role not in ("staff", "admin", "superadmin"):
            return Response({"detail": "Forbidden"}, status=403)
        fm = FamilyMember.objects.filter(id=member_id, user_id=user_id).first()
        if not fm:
            return Response({"detail": "Thành viên không tồn tại"}, status=404)
        if getattr(fm, "is_self", False):
            return Response({"detail": "Không thể xoá bản thân"}, status=400)
        fm.delete()
        return Response(status=204)
    
# ---------- Tạo booking  cho n customer  -----------
class StaffCreateAppointmentAPIView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request, user_id: int):
        role = getattr(request.user, "role", "")
        if role not in ("staff", "admin", "superadmin"):
            return Response({"detail": "Forbidden"}, status=403)
        user = CustomUser.objects.filter(id=user_id, role="customer").first()
        if not user:
            return Response({"detail": "Customer không tồn tại"}, status=404)
        in_ser = StaffBookingCreateInSerializer(data=request.data)
        in_ser.is_valid(raise_exception=True)
        data = in_ser.validated_data
        # ép owner là customer này
        if data["member"].user_id != user.id:
            return Response({"detail": "Thành viên không thuộc tài khoản này."}, status=400)
        # ---- CHẶN ĐẶT >1 LỊCH HẸN TƯƠNG LAI CHO CÙNG VẮC XIN (CÙNG MEMBER) ----
        today = timezone.now().date()
        member = data["member"]
        for item in data["items"]:
            # tuỳ StaffBookingCreateInSerializer, thường sẽ là instance Vaccine
            vaccine = item.get("vaccine") or item.get("vaccine_id")
            if not vaccine:
                continue
            if hasattr(vaccine, "id"):
                vaccine_id = vaccine.id
                vaccine_name = getattr(vaccine, "name", f"ID {vaccine.id}")
            else:
                vaccine_id = int(vaccine)
                vaccine_obj = Vaccine.objects.filter(id=vaccine_id).first()
                vaccine_name = vaccine_obj.name if vaccine_obj else f"ID {vaccine_id}"
            # tìm 1 lịch hẹn chưa tiêm (pending/confirmed) cho cùng member + vaccine, từ hôm nay trở đi
            conflict = (
                Booking.objects
                .filter(
                    user=user,
                    member=member,
                    appointment_date__gte=today,
                    status__in=["pending", "confirmed"],
                    items__vaccine_id=vaccine_id,
                )
                .order_by("appointment_date")
                .first()
            )
            if conflict:
                old_date = conflict.appointment_date.strftime("%d/%m/%Y") if conflict.appointment_date else ""
                msg = (
                    f"Đã tồn tại lịch hẹn chưa tiêm {vaccine_name} vào ngày {old_date}. "
                    "Để tránh trùng lịch, vui lòng cập nhật hoặc hủy lịch hẹn đó trước khi tạo lịch mới."
                )
                return Response({"detail": msg}, status=400)
        # map sang BookingSerializer để dùng validate() có sẵn
        payload = {
            "member_id": data["member"].id,
            "appointment_date": data["appointment_date"],
            "location": data.get("location") or "",
            "notes": data.get("notes") or "",
            "items": [{"vaccine_id": it["vaccine_id"], "quantity": it["quantity"]} for it in data["items"]],
        }
        ser = BookingSerializer(
            data=payload,
            context={"request": request, "acting_user": user}  # user ở đây là customer đã load ở trên
        )
        ser.is_valid(raise_exception=True)
        booking = ser.save()
        return Response(BookingSerializer(booking, context={"request": request}).data, status=201)
    

# --------- Đếm trước + trả danh sách chi tiết ----------
class CustomerNotificationPreviewAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        audience = request.query_params.get("audience")  # upcoming | nextdose | custom | overdue
        days_before = request.query_params.get("days_before")
        next_dose_days = request.query_params.get("next_dose_days")
        # nhận cả 2 tên
        booking_ids = (
            request.query_params.getlist("booking_ids")
            or request.query_params.getlist("customer_ids")
        )
        want_detail = request.query_params.get("detail") == "1"
        only_unscheduled = request.query_params.get("only_unscheduled") == "1"
        today = timezone.now().date()
        results = []
        count = 0
        # 1) chọn theo danh sách booking chỉ định
        if audience == "custom":
            bks = (
                Booking.objects
                .filter(id__in=booking_ids)
                .select_related("user", "member")
                .prefetch_related("items__vaccine")
            )
            count = bks.count()
            if want_detail:
                for b in bks:
                    vaccine_names = []
                    for it in b.items.all():
                        if it.vaccine:
                            vaccine_names.append(it.vaccine.name)
                    if b.vaccine:
                        vaccine_names.append(b.vaccine.name)
                    if b.package:
                        vaccine_names.append(f"Gói: {b.package.name}")

                    results.append({
                        "type": "booking",
                        "booking_id": b.id,
                        "user_id": b.user_id,
                        "user_phone": b.user.phone if b.user else "",
                        "user_email": b.user.email if b.user else "",
                        "member_name": b.member.full_name if b.member else "",
                        "appointment_date": b.appointment_date,
                        "status": b.status,
                        "vaccine": ", ".join(dict.fromkeys(vaccine_names)) if vaccine_names else "",
                    })

        # 2) lịch sắp tới
        elif audience == "upcoming":
            try:
                n = int(days_before or 0)
            except:
                n = 0
            target_date = today + timedelta(days=n)
            qs = (
                Booking.objects
                .filter(
                    appointment_date=target_date,
                    status__in=["pending", "confirmed"],
                )
                .select_related("user", "member")
                .prefetch_related("items__vaccine")
            )
            count = qs.count()
            if want_detail:
                for b in qs:
                    vaccine_names = []
                    for it in b.items.all():
                        if it.vaccine:
                            vaccine_names.append(it.vaccine.name)
                    if b.vaccine:
                        vaccine_names.append(b.vaccine.name)
                    if b.package:
                        vaccine_names.append(f"Gói: {b.package.name}")

                    results.append({
                        "type": "booking",
                        "booking_id": b.id,
                        "user_id": b.user_id,
                        "user_phone": b.user.phone if b.user else "",
                        "user_email": b.user.email if b.user else "",
                        "member_name": b.member.full_name if b.member else "",
                        "appointment_date": b.appointment_date,
                        "status": b.status,
                        "vaccine": ", ".join(dict.fromkeys(vaccine_names)) if vaccine_names else "",
                    })
        # 3) mũi tiếp theo từ sổ tiêm (theo phác đồ trong sổ)
        elif audience == "nextdose":
            try:
                n = int(next_dose_days or 3)
            except Exception:
                n = 3
            to = today + timedelta(days=n)
            # Lấy tất cả record có next_dose_date trong khoảng [today, today + n]
            recs = (
                VaccinationRecord.objects
                .filter(
                    next_dose_date__gte=today,
                    next_dose_date__lte=to,
                )
                .select_related("family_member__user", "vaccine", "disease")
            )
            # Nếu chỉ muốn nhắc các mũi CHƯA có lịch hẹn (only_unscheduled=1)
            bookings_by_key = {}
            if only_unscheduled:
                member_ids = [r.family_member_id for r in recs if r.family_member_id]
                if member_ids:
                    bks = (
                        Booking.objects.filter(
                            member_id__in=member_ids,
                            appointment_date__gte=today,
                            appointment_date__lte=to,
                        )
                        .exclude(status__in=["cancelled"])
                        .select_related("member", "user")
                        .prefetch_related("items__vaccine")
                    )
                    for b in bks:
                        if b.items.exists():
                            for it in b.items.all():
                                bookings_by_key[
                                    (b.member_id, b.appointment_date, it.vaccine_id)
                                ] = True
                        else:
                            bookings_by_key[
                                (b.member_id, b.appointment_date, b.vaccine_id)
                            ] = True
            results = []
            for r in recs:
                fm = r.family_member
                usr = fm.user if fm else None
                if not usr:
                    continue
                v = r.vaccine
                # Nếu đã có booking tương ứng cho mũi này trong khoảng thì bỏ qua
                if only_unscheduled and v:
                    key = (
                        fm.id if fm else None,
                        r.next_dose_date,
                        v.id,
                    )
                    if bookings_by_key.get(key):
                        continue
                vaccine_name = v.name if v else (r.vaccine_name or "")
                disease_name = (
                    r.disease.name
                    if r.disease
                    else (
                        v.disease.name
                        if v and v.disease
                        else ""
                    )
                )
                interval = getattr(v, "interval_days", None) if v else None
                total_doses = getattr(v, "doses_required", None) if v else None
                results.append({
                    "type": "record",
                    "record_id": r.id,
                    "user_id": usr.id if usr else None,
                    "user_phone": usr.phone if usr else "",
                    "user_email": usr.email if usr else "",
                    "member_name": fm.full_name if fm else "",
                    "next_dose_date": r.next_dose_date,
                    "status": "nextdose",
                    "vaccine": vaccine_name,
                    "disease_name": disease_name,
                    "interval": interval,
                    "total_doses": total_doses,
                })

            count = len(results)
        # 4) trễ hẹn
        elif audience == "overdue":
            qs = (
                Booking.objects
                .filter(appointment_date__lt=today)
                .exclude(status__in=["completed", "cancelled"])
                .select_related("user", "member")
                .prefetch_related("items__vaccine")
            )
            count = qs.count()
            if want_detail:
                for b in qs:
                    vaccine_names = []
                    for it in b.items.all():
                        if it.vaccine:
                            vaccine_names.append(it.vaccine.name)
                    if b.vaccine:
                        vaccine_names.append(b.vaccine.name)
                    if b.package:
                        vaccine_names.append(f"Gói: {b.package.name}")
                    results.append({
                        "type": "booking",
                        "booking_id": b.id,
                        "user_id": b.user_id,
                        "user_phone": b.user.phone if b.user else "",
                        "user_email": b.user.email if b.user else "",
                        "member_name": b.member.full_name if b.member else "",
                        "appointment_date": b.appointment_date,
                        "status": "overdue",
                        "vaccine": ", ".join(dict.fromkeys(vaccine_names)) if vaccine_names else "",
                    })
        return Response({"count": count, "results": results if want_detail else []})


def format_vi_date(val):
    if not val:
        return ""
    if isinstance(val, (datetime, date_cls)):
        d = val.date() if isinstance(val, datetime) else val
        return f"{d.day:02d}/{d.month:02d}/{d.year}"
    s = str(val)
    try:
        d = datetime.fromisoformat(s).date()
        return f"{d.day:02d}/{d.month:02d}/{d.year}"
    except Exception:
        pass
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        try:
            y, m, d = s.split("-")
            return f"{int(d):02d}/{int(m):02d}/{int(y)}"
        except Exception:
            pass
    return s

def render_msg(tpl: str, ctx: dict | None) -> str:
    if not tpl:
        return ""
    ctx = ctx or {}

    def repl(m):
        key = m.group(1).strip()
        val = ctx.get(key, "")
        if key in ("date", "appointment_date"):
            return format_vi_date(val)
        return "" if val is None else str(val)

    return re.sub(r"\{\{([^}]+)\}\}", repl, tpl)

def send_notification_email(to_email: str, subject: str, body: str):
    if not to_email:
        return

    main_from = "#1186f3"  # xanh dương
    main_to   = "#1af5f5"  # xanh ngọc

    safe_subject = escape(subject or "").strip() or "Thông báo lịch tiêm"
    safe_body = escape(body or "")
    body_html = safe_body.replace("\n", "<br>")

    year = datetime.now().year

    app_url = getattr(
        settings,
        "EVACCINE_APP_URL",
        "http://localhost:3000/notifications"
    )

    # Gửi mail nhắc lịch tự động
    html_body = f"""\
    <!DOCTYPE html>
    <html lang="vi">
    <head>
        <meta charset="UTF-8" />
        <title>{safe_subject}</title>
    </head>
    <body style="margin:0;padding:0;background-color:#f3f4f6;font-family:system-ui,-apple-system,Segoe UI,Roboto,sans-serif;">
        <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#f3f4f6;padding:24px 0;">
        <tr>
            <td align="center">
            <table width="640" cellspacing="0" cellpadding="0" style="background:#ffffff;border-radius:18px;overflow:hidden;box-shadow:0 20px 45px rgba(15,23,42,0.16);">
                <!-- Header -->
                <tr>
                <td style="padding:20px 24px;background:linear-gradient(120deg,{main_from},{main_to});color:#ffffff;">
                    <table width="100%" cellspacing="0" cellpadding="0">
                    <tr>
                        <td style="font-size:22px;font-weight:800;letter-spacing:0.03em; white-space:nowrap;">
                            <img src="cid:evaccine-logo"
                                 alt="E-VACCINE" width="32" height="32"
                                 style="vertical-align:middle;border-radius:50%;border:2px solid #fff;margin-right:8px;">
                            E-VACCINE
                        </td>
                        <td align="right" style="font-size:13px;opacity:0.85;">
                            <strong>Hệ thống tiêm chủng thông minh</strong>
                        </td>
                    </tr>
                    </table>
                </td>
                </tr>
                <!-- Hero title -->
                <tr>
                <td style="padding:20px 24px 8px 24px;">
                    <div style="font-size:18px;font-weight:700;color:#0f172a;margin-bottom:4px;">
                    {safe_subject}
                    </div>
                    <div style="font-size:13px;color:#6b7280;">
                    Đây là email nhắc lịch tự động từ hệ thống E-Vaccine.
                    </div>
                </td>
                </tr>
                <!-- Appointment card -->
                <tr>
                <td style="padding:0 24px 16px 24px;">
                    <table width="100%" cellspacing="0" cellpadding="0" style="border-radius:14px;background:linear-gradient(135deg,#eff6ff,#ecfeff);border:1px solid #dbeafe;">
                    <tr>
                        <td style="padding:14px 16px 4px 16px;">
                        <div style="font-size:13px;font-weight:600;color:#1d4ed8;margin-bottom:6px;">
                            Thông tin lịch tiêm
                        </div>
                        <div style="font-size:14px;color:#111827;line-height:1.7;">
                            {body_html}
                        </div>
                        </td>
                    </tr>
                    <tr>
                        <td style="padding:8px 16px 14px 16px;">
                        <a href="{app_url}"
                            style="display:inline-block;padding:9px 18px;border-radius:999px;background:#0ea5e9;color:#ffffff;
                                    font-size:13px;font-weight:600;text-decoration:none;">
                                Xem chi tiết lịch hẹn
                        </a>
                        <span style="font-size:11px;color:#6b7280;margin-left:8px;">
                            (Đăng nhập tài khoản E-Vaccine để quản lý lịch tiêm)
                        </span>
                        </td>
                    </tr>
                    </table>
                </td>
                </tr>
                <!-- Tips -->
                <tr>
                <td style="padding:0 24px 20px 24px;">
                    <table width="100%" cellspacing="0" cellpadding="0" style="border-radius:12px;background:#fefce8;border:1px solid #facc15;">
                    <tr>
                        <td style="padding:10px 14px;font-size:12px;color:#854d0e;line-height:1.6;">
                        <strong>Lưu ý khi đi tiêm:</strong>
                        <ul style="margin:6px 0 0 16px;padding:0;">
                            <li>Đem theo CCCD/CMND hoặc giấy tờ tùy thân.</li>
                            <li>Thông báo cho nhân viên y tế nếu bạn đang mang thai, dùng thuốc, hoặc có bệnh lý nền.</li>
                            <li>Đến sớm 10-15 phút để được hướng dẫn làm thủ tục.</li>
                        </ul>
                        </td>
                    </tr>
                    </table>
                </td>
                </tr>
                <!-- Footer -->
                <tr>
                <td style="padding:12px 24px 16px 24px;border-top:1px solid #e5e7eb;text-align:center;font-size:11px;color:#9ca3af;">
                    Email được gửi từ hệ thống E-Vaccine, vui lòng không trả lời trực tiếp email này.<br/>
                    © {year} Trung tâm tiêm chủng E-Vaccine. Mọi quyền lợi đều được ưu tiên.
                </td>
                </tr>
            </table>
            </td>
        </tr>
        </table>
    </body>
    </html>
    """
    plain_text = strip_tags(html_body)
    # Tạo email (giống ForgotPasswordAPIView)
    email = EmailMultiAlternatives(
        subject=safe_subject,
        body=plain_text,
        from_email=settings.EMAIL_HOST_USER,
        to=[to_email],
    )
    email.attach_alternative(html_body, "text/html")
    # Gắn logo từ static/images/logo.jpg (bạn đổi tên file cho đúng)
    logo_path = Path(settings.BASE_DIR) / "static" / "images" / "logo.jpg"
    if logo_path.exists():
        with open(logo_path, "rb") as f:
            img_logo = MIMEImage(f.read())
            img_logo.add_header("Content-ID", "<evaccine-logo>")
            img_logo.add_header("Content-Disposition", "inline", filename="logo.jpg")
            email.attach(img_logo)

    email.send(fail_silently=False)
    
class MyNotificationsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = CustomerNotification.objects.filter(user=request.user).order_by("-created_at")
        data = []
        for n in qs:
            item = CustomerNotificationSerializer(n).data
            item["sent_at"] = n.created_at
            msg = item.get("message") or ""
            if "{{" in msg and "}}" in msg:
                ctx = {
                    "name": request.user.full_name or request.user.email or request.user.phone or "",
                    # meta từ lúc gửi
                    "date": (n.meta or {}).get("appointment_date") or "",
                    "appointment_date": (n.meta or {}).get("appointment_date") or "",
                    "member": (n.meta or {}).get("member_name") or "",
                    "vaccine": ", ".join((n.meta or {}).get("vaccines") or []),
                    "location": (n.meta or {}).get("location") or "",
                }
                item["message"] = render_msg(msg, ctx)

            data.append(item)
        return Response(data)
class MyNotificationMarkReadAPIView(APIView):
    permission_classes = [permissions.IsAuthenticated]
    def post(self, request, pk):
        notif = CustomerNotification.objects.filter(id=pk, user=request.user).first()
        if not notif:
            return Response(status=404)
        notif.is_read = True
        notif.save(update_fields=["is_read"])
        return Response({"detail": "OK"}, status=200)
    
    
    
    